# etl_sigrid/application/steps/load_excel_aux_step.py
"""
Step que resuelve, obtiene y valida los tres Excels auxiliares de Negocio.

Cada `AUX_EXCEL_*` puede ser una ruta del sistema de ficheros o una URI de
Azure Blob Storage: el step no lo sabe ni le importa: pide el contenido al
puerto `AuxFileSource` y trabaja siempre en memoria, que es lo que permite que
el ETL corra en un contenedor sin sistema de ficheros propio.

FRONTERA de F-004: aquí se LEE y se VALIDA, no se carga nada a `aux.*`. Las
tablas destino no existen y el esquema de los tres libros no está en el
repositorio; inventarlo sería inventar el modelo de datos de Negocio. La carga
es feature propia (decisión abierta DA-1 en progress/current.md).

Desenlaces:
  - SKIPPED  : ninguna variable configurada (el caso normal en un clon nuevo).
  - SUCCESS  : todo lo configurado se obtuvo y abre como libro de Excel.
  - FAILED   : alguno falló. Se reportan TODOS en un único mensaje, no el
               primero: nadie quiere arreglar de uno en uno esperando al job
               nocturno entre intento e intento.
"""

from __future__ import annotations

from collections.abc import Callable
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from config.settings import Settings
from etl_sigrid.application.steps.base import PipelineStep
from etl_sigrid.domain.entities import StepResult, StepStatus
from etl_sigrid.infrastructure.excel.aux_file_source import (
    AuxFileError,
    AuxFileRef,
    AuxFileSource,
    get_aux_file_source,
    parse_aux_file_ref,
)
from etl_sigrid.infrastructure.logging_config import get_logger

logger = get_logger(__name__)


class LoadExcelAuxStep(PipelineStep):
    """Obtiene TipoPartida, TipoCoste y mapeo_proporcionales, y comprueba que abren."""

    def __init__(
        self,
        settings: Settings,
        *,
        source_factory: Callable[[AuxFileRef], AuxFileSource] = get_aux_file_source,
    ) -> None:
        """
        `source_factory` es el punto de inyección: en producción resuelve el
        adaptador por el origen de la referencia; en los tests devuelve un doble
        y así ningún test toca red ni disco.
        """
        self._settings = settings
        self._source_factory = source_factory

    @property
    def name(self) -> str:
        return "load_excel_aux"

    @property
    def stage(self) -> str:
        return "load_aux"

    def run(self) -> StepResult:
        result = self._new_result()
        entradas = self._settings.aux_excel.entries()

        configuradas = [(n, var, valor) for n, var, valor in entradas if valor.strip()]
        omitidos = [n for n, _, valor in entradas if not valor.strip()]

        if not configuradas:
            faltantes = ", ".join(var for _, var, _ in entradas)
            result.status = StepStatus.SKIPPED
            result.error_message = (
                f"Ningún Excel auxiliar configurado ({faltantes} están vacías): "
                f"no hay nada que leer. Configura una ruta local o una URI de blob "
                f"si quieres que este paso haga algo."
            )
            result.metadata = {"files": {}, "omitidos": omitidos}
            result.finished_at = datetime.utcnow()
            logger.info("aux_files_skipped", omitidos=omitidos)
            return result

        ficheros: dict[str, dict[str, Any]] = {}
        errores: list[str] = []

        for logical_name, env_var, valor in configuradas:
            try:
                ficheros[logical_name] = self._leer_uno(logical_name, env_var, valor)
            except AuxFileError as exc:
                errores.append(str(exc))
            except Exception as exc:  # noqa: BLE001 - el fallo de uno no tumba a los demás
                # Red de seguridad: cualquier fallo no previsto de la fuente se
                # convierte igualmente en un error atribuido a SU fichero, para
                # que el mensaje final siga listándolos todos.
                errores.append(
                    f"Fallo inesperado con el Excel auxiliar '{logical_name}' "
                    f"(variable {env_var}): {type(exc).__name__}: {exc}"
                )

        result.metadata = {"files": ficheros, "omitidos": omitidos}
        result.rows_processed = len(ficheros)
        result.finished_at = datetime.utcnow()

        if errores:
            unidos = "\n\n  · " + "\n\n  · ".join(errores)
            result.status = StepStatus.FAILED
            result.error_message = (
                f"{len(errores)} Excel(s) auxiliar(es) no se pudieron leer:{unidos}"
            )
            logger.error("aux_files_failed", fallos=len(errores), leidos=len(ficheros))
            return result

        result.status = StepStatus.SUCCESS
        logger.info("aux_files_done", leidos=len(ficheros), omitidos=omitidos)
        return result

    # ------------------------------------------------------------------
    # Un fichero: resolver, obtener y validar
    # ------------------------------------------------------------------

    def _leer_uno(self, logical_name: str, env_var: str, valor: str) -> dict[str, Any]:
        """Devuelve la ficha del fichero (origen, ubicación, tamaño y hojas)."""
        ref = parse_aux_file_ref(logical_name, env_var, valor)
        datos = self._source_factory(ref).read_bytes(ref)

        try:
            libro = load_workbook(BytesIO(datos), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001 - openpyxl lanza de todo ante un fichero roto
            raise AuxFileError(
                f"El Excel auxiliar '{logical_name}' se obtuvo de {ref.display} "
                f"(variable {env_var}, {len(datos)} bytes) pero no abre como libro de "
                f"Excel: {type(exc).__name__}: {exc}. Comprueba que es un .xlsx válido, "
                f"no está corrupto y no es un .xls antiguo."
            ) from exc

        try:
            hojas = list(libro.sheetnames)
        finally:
            libro.close()

        logger.info(
            "aux_file_read",
            logical_name=logical_name,
            origen=ref.origin,
            ubicacion=ref.display,
            bytes=len(datos),
            hojas=len(hojas),
        )
        return {
            "origen": ref.origin,
            "ubicacion": ref.display,
            "bytes": len(datos),
            "hojas": hojas,
        }
